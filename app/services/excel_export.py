import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.utils.config import get_settings
from app.utils.logging import logger

settings = get_settings()

COLUMNS_ORDER = [
    "sr_no", "kas_name", "customer_name", "site_location", "country",
    "incoterm", "direct_sales_wh_movement", "po_forecast",
    "category", "sub_category", "cust_part_no", "maini_part_no",
    "open_qty", "unit_price", "currency", "unit_price_inr",
    "total_inr", "doc_date", "ship_date", "sales_month",
]
DISPLAY_HEADERS = {
    "sr_no": "S No",
    "kas_name": "KAS Name",
    "customer_name": "Customer Name",
    "site_location": "Site Location",
    "country": "Country",
    "incoterm": "Incoterm",
    "direct_sales_wh_movement": "Direct Sales / WH Movement",
    "po_forecast": "PO # / Forecast",
    "category": "Category",
    "sub_category": "Sub Category",
    "cust_part_no": "Cust Part #",
    "maini_part_no": "Maini Part #",
    "open_qty": "Open Qty",
    "unit_price": "Unit Price",
    "currency": "Currency",
    "unit_price_inr": "Unit Price in INR",
    "total_inr": "Total in INR",
    "doc_date": "Doc Date",
    "ship_date": "Ship Date",
    "sales_month": "Sales Month",
}


def _is_forecast_label(po_value) -> bool:
    v = str(po_value or "").strip()
    return bool(v) and not v.replace("-", "").replace("/", "").isdigit()


def _month_label(date_str: str) -> str:
    if not date_str:
        return ""
    try:
        from dateutil import parser as dp
        return dp.parse(str(date_str)).strftime("%b-%Y")
    except Exception:
        return ""


def _expand_forecast_rows(items: list[dict]) -> list[dict]:
    """Expand forecast rows that carry a forecast_schedule into one row per
    date+quantity entry so the full schedule appears in the exported file.
    Non-forecast rows and forecast rows without schedule data pass through unchanged.
    """
    expanded: list[dict] = []
    sr = 0

    for item in items:
        po = str(item.get("po_forecast", "") or "").strip()
        schedule: dict = item.get("forecast_schedule") or {}

        if _is_forecast_label(po) and schedule:
            unit_price = float(item.get("unit_price") or 0)
            for date_str in sorted(schedule.keys()):
                qty = float(schedule[date_str])
                if qty <= 0:
                    continue
                sr += 1
                row = {k: v for k, v in item.items() if k != "forecast_schedule"}
                row["sr_no"] = sr
                row["ship_date"] = date_str
                row["open_qty"] = qty
                row["sales_month"] = _month_label(date_str)
                row["unit_price_inr"] = round(unit_price, 2)
                row["total_inr"] = round(qty * unit_price, 2)
                expanded.append(row)
        else:
            sr += 1
            row = {k: v for k, v in item.items() if k != "forecast_schedule"}
            row["sr_no"] = sr
            expanded.append(row)

    return expanded


def export_zso_to_excel(zso_data: dict, output_dir: str | None = None) -> str:
    """Generate a formatted Excel file from ZSO report data.

    Forecast rows that carry a forecast_schedule are expanded into one row
    per date+quantity so the full schedule is visible in the export.
    """
    output_dir = output_dir or os.path.join(settings.UPLOAD_DIR, "exports")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ZSO_Report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    raw_items = zso_data.get("items", [])
    if not raw_items:
        logger.warning("No items to export")
        return ""

    # Expand forecast rows → one row per schedule date+qty
    items = _expand_forecast_rows(raw_items)
    forecast_expanded = len(items) - len(raw_items)
    if forecast_expanded > 0:
        logger.info(f"Export: expanded forecast rows → {len(items)} total rows (was {len(raw_items)})")

    df = pd.DataFrame(items)
    df = df[[c for c in COLUMNS_ORDER if c in df.columns]]
    df.rename(columns=DISPLAY_HEADERS, inplace=True)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Summary sheet — reflect the expanded row count
        summary_data = {
            "Field": ["KAS Name", "Generated At", "Total Rows (exported)", "Matched Items", "Total INR", "Status"],
            "Value": [
                zso_data.get("kas_name", ""),
                zso_data.get("generated_at", ""),
                len(items),          # expanded count
                zso_data.get("matched_items", 0),
                zso_data.get("total_inr", 0),
                "Generated",
            ],
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        # Detail sheet
        df.to_excel(writer, sheet_name="ZSO Details", index=False)

        # Style the workbook
        wb = writer.book
        _style_sheet(wb["Summary"], header_color="1F4E79")
        _style_sheet(wb["ZSO Details"], header_color="FFFF00", header_font_color="000000")

    logger.info(f"ZSO Excel exported: {filepath}")
    return filepath


def _style_sheet(ws, header_color: str = "2E75B6", header_font_color: str = "FFFFFF") -> None:
    header_font = Font(bold=True, color=header_font_color, size=11)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)
